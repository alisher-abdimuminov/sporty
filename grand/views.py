import requests
import os
import openpyxl
import datetime

from django.views import View
from django.http import HttpRequest, JsonResponse, HttpResponse, HttpResponseRedirect
from django.shortcuts import render, redirect, get_object_or_404
from .client import oAuth2Client
from django.conf import settings
from .models import Student, Criteria, StudentFiles
from .forms import StudentFilesForm, StudentFileForm

from django.core.files.base import ContentFile
from urllib.parse import urlparse
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import user_passes_test

from openpyxl.utils import get_column_letter

from .serializers import StudentSerializer, StudentFilesSerializer


date_string = "2025-09-5"
date_object = datetime.datetime.strptime(date_string, "%Y-%m-%d")
current_datetime = datetime.datetime.now()


def landing_page(request):
    return render(request, "homepage.html")


def export_social_activity_excel(request, pk):
    student = get_object_or_404(Student, pk=pk)
    student_files = StudentFiles.objects.filter(student=student)
    criteria_list = Criteria.objects.all()

    # Yangi Excel fayl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Social Activity Scores"

    # Headerlar: Talaba, Fakultet, Guruh, Mezonlar, Umumiy ball
    headers = ["Talaba", "Fakultet", "Guruh"]
    headers += [c.title for c in criteria_list]
    headers.append("Umumiy ball")
    ws.append(headers)

    # Talaba haqida ma'lumot
    row = [
        student.student_name,
        student.faculty if student.faculty else "",
        student.groups[0].get("name", "-") if student.groups else "",
    ]

    total_score = 0
    for criterion in criteria_list:
        file = student_files.filter(criteria=criterion).first()
        score = file.task_score if file and file.task_score else 0
        row.append(score)
        total_score += score

    row.append(total_score)
    ws.append(row)
    for i, column in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 20

    # Excel response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{student.student_name}_social_scores.xlsx"'
    )
    wb.save(response)
    return response


@user_passes_test(lambda u: u.is_superuser)
def reset_score(request):
    update_count = StudentFiles.objects.update(is_scored=False, task_score=0)
    return JsonResponse({"message": f"{update_count} ta student yangilandi"})


def home(request: HttpRequest):
    student_hemis_id = request.COOKIES.get("student_hemis_id")
    if not student_hemis_id:
        return redirect("/auth/")
    student = Student.objects.get(student_id_number=student_hemis_id)
    return render(request, "index.html", context={"student": student})


def faculty_list(request):
    faculties = [
        "Tabiiy fanlar",
        "Aniq va amaliy fanlar",
        "Maktabgacha va Boshlang'ich ta'lim",
        "Pedagogika va jismoniy madaniyat",
        "Tillar",
        "Magistratura mutaxassisligi",
    ]
    return render(request, "faculty_list.html", {"faculties": faculties})


def supervisor_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect("faculty_list")
        else:
            messages.error(request, "Invalid username or password.")

    return render(request, "login.html")


def faculty_detail(request, name):
    students = Student.objects.filter(faculty=name)

    serializer = StudentSerializer(students, many=True)
    return render(
        request,
        "faculty-students.html",
        {"students": serializer.data, "faculty_name": name},
    )


def student_files(request, pk):
    try:
        student = Student.objects.get(student_id_number=pk)
        serializer_student = StudentSerializer(student)
        print(serializer_student.data)
    except Student.DoesNotExist:
        return HttpResponse("Student topilmadi", status=404)

    student_fl = StudentFiles.objects.filter(student=student)
    serializer = StudentFilesSerializer(student_fl, many=True)

    return render(
        request,
        "student_files.html",
        {"students_fl": serializer.data, "student": serializer_student.data},
    )


def score_file(request, pk):
    file = get_object_or_404(StudentFiles, pk=pk)
    student = Student.objects.get(id=file.student_id)
    serializer = StudentSerializer(student)

    if request.method == "POST":
        task_score = int(request.POST.get("task_score"))
        comment = request.POST.get("supervisor_comment", "")
        file.task_score = task_score
        file.supervisor_comment = comment
        file.is_scored = True
        file.supervisor = (
            request.user.supervisor
        )  # Assuming supervisor is logged in user
        file.save()
        return redirect("student_files", pk=file.student.student_id_number)

    return render(
        request, "score_file.html", {"student_file": file, "student": serializer.data}
    )


def student_settings(request):
    pass


def logout_view(request: HttpRequest):
    response = redirect("/")
    response.delete_cookie("student_hemis_id")
    response.delete_cookie("hemis_access_token")

    return response


def criteria(request: HttpRequest):
    if request.method == "GET":
        student_hemis_id = request.COOKIES.get("student_hemis_id")
        if not student_hemis_id:
            return redirect("/auth/")
        criterias = Criteria.objects.all()
        return render(request, "criteria.html", {"criterias": criterias})
    return render(request, "criteria.htmlt")


def upload_file(request: HttpRequest, criteria_id):
    criteria = get_object_or_404(Criteria, pk=criteria_id)
    # serializer =
    print(criteria)
    student_hemis_id = request.COOKIES.get("student_hemis_id")
    if not student_hemis_id:
        return redirect("/auth/")
    student = Student.objects.get(student_id_number=student_hemis_id)

    if request.method == "POST":
        form = StudentFilesForm(
            request.POST or None,
            request.FILES or None,
            student=student,
            criteria=criteria,
        )

        if form.is_valid():
            student_file = form.save(commit=False)
            student_file.criteria = criteria
            student_file.student = student
            student_file.save()
            return redirect("/criteria/")
    else:
        form = StudentFilesForm()

    return render(request, "upload_file.html", {"form": form, "criteria": criteria})


def student_profile(request: HttpRequest, student_id):
    student_hemis_id = request.COOKIES.get("student_hemis_id")

    if not student_hemis_id:
        return redirect("/auth/")

    # student_id argumentini ishlating, lekin siz cache dan ham olayapsiz,
    # agar kerak bo'lsa, student_id ni tekshiring
    student = get_object_or_404(Student, student_id_number=student_id)

    files = StudentFiles.objects.filter(student=student)
    if request.method == "POST":
        file_id = request.POST.get("file_id")
        action = request.POST.get("action")

        file_obj = get_object_or_404(StudentFiles, id=file_id, student=student)

        if action == "edit":
            form = StudentFileForm(request.POST, request.FILES, instance=file_obj)
            if form.is_valid():
                form.save()
                return redirect("profile", student_id=student.student_id_number)

        elif action == "delete":
            file_obj.delete()
            return redirect("profile", student_id=student.student_id_number)

    return render(
        request,
        "profile.html",
        {
            "student": student,
            "files": files,
            "form": StudentFileForm(),
        },
    )


def contact(request: HttpRequest):
    student_hemis_id = request.COOKIES.get("student_hemis_id")
    if not student_hemis_id:
        return redirect("/auth/")
    return render(request, "help-center.html")


def download_image_from_url(url):
    response = requests.get(url)
    if response.status_code == 200:
        parsed_url = urlparse(url)
        filename = os.path.basename(parsed_url.path)
        return ContentFile(response.content, name=filename)
    return None


class AuthLoginView(View):
    def get(self, request):
        client = oAuth2Client(
            client_id=settings.CLIENT_ID,
            client_secret=settings.CLIENT_SECRET,
            redirect_uri=settings.REDIRECT_URI,
            authorize_url=settings.AUTHORIZE_URL,
            token_url=settings.ACCESS_TOKEN_URL,
            resource_owner_url=settings.RESOURCE_OWNER_URL,
        )
        authorization_url = client.get_authorization_url()

        return HttpResponseRedirect(authorization_url)


class AuthCallbackView(View):
    def get(self, request):

        code = request.GET.get("code")
        if code is None:
            return JsonResponse({"error": "code is missing!"})

        client = oAuth2Client(
            client_id=settings.CLIENT_ID,
            client_secret=settings.CLIENT_SECRET,
            redirect_uri=settings.REDIRECT_URI,
            authorize_url=settings.AUTHORIZE_URL,
            token_url=settings.ACCESS_TOKEN_URL,
            resource_owner_url=settings.RESOURCE_OWNER_URL,
        )
        access_token_response = client.get_access_token(code)

        full_info = {}
        if "access_token" in access_token_response:
            access_token = access_token_response["access_token"]
            user_details = client.get_user_details(access_token)
            full_info["details"] = user_details
            full_info["token"] = access_token
            student_gpa = user_details["data"]["avg_gpa"]
            student_id = user_details["data"]["student_id_number"]

            student = Student.objects.filter(student_id_number=student_id)

            if not student.exists() and date_object < current_datetime:
                messages.error(request, "Tizimda roʻyxatdan oʻtish tugallandi.")
                return redirect("/")

            if float(student_gpa) < 3.50:
                messages.error(
                    request,
                    "Sizning GPA balingiz yetarli emas. Kamida 3.5 bo‘lishi kerak.",
                )
                return redirect("/")

            # cache.set('hemis_access_token', access_token, timeout=1800)
            # cache.set('student_hemis_id', user_details['student_id_number'], timeout=1800)

            if not Student.objects.filter(
                student_id_number=user_details["student_id_number"]
            ).exists():
                student = Student.objects.create(
                    student_name=user_details["name"],
                    phone_number=user_details["phone"],
                    student_id_number=user_details["student_id_number"],
                    email=user_details["email"],
                    passport_number=user_details["passport_number"],
                    birth_date=user_details["birth_date"],
                    groups=user_details.get("groups", []),
                    studentStatus=user_details["data"]["studentStatus"]["name"],
                    paymentForm=user_details["data"]["paymentForm"]["name"],
                    faculty=user_details["data"]["faculty"]["name"],
                    level=user_details["data"]["level"]["name"],
                    avg_gpa=user_details["data"]["avg_gpa"],
                )
                image_file = download_image_from_url(user_details["picture"])
                if image_file:
                    student.student_imeg.save(image_file.name, image_file, save=False)
                    student.save()
            response = redirect("/home/")
            response.set_cookie(
                "student_hemis_id", user_details["student_id_number"], max_age=1800
            )
            response.set_cookie("hemis_access_token", access_token, max_age=1800)
            return response
        else:
            return JsonResponse(
                {"status": False, "error": "Failed to obtain access token"}, status=400
            )


def error_404_view(request, exception):
    return render(request, "404.html")


# def error_500_view(request):
#     return render(request, '404.html', status=500)
